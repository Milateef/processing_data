from openpyxl import Workbook
import os
import re
current_dir = os.getcwd()
file_list = os.listdir(current_dir)

wb = Workbook()

for source_file_name in file_list:
    if '.txt' in source_file_name:
        with open(source_file_name, 'r') as sf:
            ws = wb.create_sheet(source_file_name)

            row_num = 1
            for line_of_sf in sf:
                line_split_list = re.split(',|\t', line_of_sf.rstrip())
                column_num = 1
                for cell_value in line_split_list:
                    ws.cell(row=row_num, column=column_num).value = cell_value
                    column_num = column_num + 1
                row_num = row_num + 1

        sf.close()

del_sheet = wb.get_sheet_by_name('Sheet')
wb.remove_sheet(del_sheet)
wb.save('result.xlsx')
