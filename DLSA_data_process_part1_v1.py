from openpyxl import Workbook   # import module
from openpyxl.styles import PatternFill
import os
import time

print("This process takes a couple of seconds... ...")

current_dir = os.getcwd()   # get current directory
file_list = os.listdir(current_dir)  # list file under dir

SDE_header = []     # store SDE=0x**()... for total dut
VDD_header = []     # store VDD=0x**()...
VDDSA_header = []   # store VDDSA=0x**()...
VDD_flag = 0        # represent this file is VDD vs SDE
VDDSA_flag = 0      # represent this file is VDDSA vs SDE
VCC_flag = 0        # when encouner VCC, then insert a row. for total dut file

SDE_DAC = []        # store SDE shmoo value
VDD_DAC = []        # store VDD shmoo value
VDDSA_DAC = []      # store VDDSA shmoo value
SDE_shmoo_step = 2  # SDE shmoo step
# when the value = 1, then find the default SDE and VDD or VDDSA and highlight
pattern_flag = 0

SDE_default_list = []       # store SDE default value of each dut
VDD_default_list = []       # store VDD default value of each dut
VDDSA_default_list = []     # store VDDSA default value of each dut

for i in range(0x3F, 0x00, -SDE_shmoo_step):
    SDE_DAC.append(i)       # generate SDE_DAC

for i in range(0xF, 0x00, -1):
    VDD_DAC.append(i)
    VDDSA_DAC.append(i)     # generate VDD_DAC and VDDSA_DAC

# SDE_DAC.append(0x00)      # don't need this
VDD_DAC.append(0x00)
VDDSA_DAC.append(0x00)
# fill cell with red
redFill = PatternFill(
    start_color='00FF9999',
    end_color='00FF9999',
    fill_type='solid'
)
# fill cell with green
greenFill = PatternFill(
    start_color='0099FF99',
    end_color='0099FF99',
    fill_type='solid'
)
# fill cell with yellow
yellowFill = PatternFill(
    start_color='00FFFF33',
    end_color='00FFFF33',
    fill_type='solid'
)


# get index of a sub string, reture -1 if when whole string didn't include sub string
def get_index(whole, sub):
    try:
        index = whole.index(sub)
        return index
    except Exception:
        return -1


# invert str to num, both intger and hexdecimal. return -1 when the str can't be inverted
def str_to_num(original_str):
    try:
        inverted_num = int(original_str)
    except Exception:
        inverted_num = -1

    if inverted_num == -1:
        try:
            inverted_num = int(original_str, 16)
        except Exception:
            inverted_num = -1
    return inverted_num


wb = Workbook()         # create workbook

for source_file_name in file_list:      # loop the files in current dir
    # if the file is for 1 DUT
    if 'tb_DLSA_Test_SDE' in source_file_name and 'DUT' in source_file_name:
        with open(source_file_name, 'r') as sf:
            # start_index and end_index are used to generate the sheet name
            start_index = source_file_name.index('SDE')

            if get_index(source_file_name, 'DUT') != -1:
                end_index = get_index(source_file_name, 'DUT')
                sheet_name = source_file_name[start_index: end_index + 4]

            ws = wb.create_sheet(sheet_name)
            # get the default values of SDE, VDD, VDDSA
            default_values = sf.readline().rstrip().split(',')
            for value in default_values:

                if 'SDE' in value:
                    SDE_default_list.append(value.split('=')[-1])
                    try_SDE_index = get_index(
                        SDE_DAC, int(SDE_default_list[-1], 16))

                    if try_SDE_index != -1:
                        def_SDE_index = try_SDE_index

                    else:
                        def_SDE_index = get_index(
                            SDE_DAC, int(SDE_default_list[-1], 16) - 1)

                elif 'VDD' in value and 'VDDSA' not in value:
                    VDD_default_list.append(value.split('=')[-1])
                    def_VDD_index = get_index(
                        VDD_DAC, int(VDD_default_list[-1], 16))

                elif 'VDDSA' in value:
                    VDDSA_default_list.append(value.split('=')[-1])
                    def_VDDSA_index = get_index(
                        VDDSA_DAC, int(VDDSA_default_list[-1], 16))

            row_num = 1     # count the row number

            for line_of_sf in sf:
                # when encounter 'Pattern', then can find the default value and fill yellow
                if 'Pattern' in line_of_sf:
                    pattern = line_of_sf.split(' ')[1]
                    pattern_flag = 1

                line_split_list = line_of_sf.rstrip().split(',')

                column_num = 1
                # use SDE_header, VDD_header, VDDSA_header to store the headers, then can fill these headers to all DUT file
                for cell_value in line_split_list:

                    if 'SDE' in cell_value and cell_value not in SDE_header:
                        SDE_header.append(cell_value)

                    if 'VDDSA' in cell_value:
                        VDDSA_flag = 1

                    elif 'VDD' in cell_value:
                        VDD_flag = 1

                    if VDDSA_flag == 1 and cell_value not in VDDSA_header:
                        VDDSA_header.append(cell_value)

                    elif VDD_flag == 1 and cell_value not in VDD_header:
                        VDD_header.append(cell_value)
                    # if the value is number, then we can judge the color should be red or green or yellow
                    cell_num_value = str_to_num(cell_value)

                    if cell_num_value != -1:
                        ws.cell(row=row_num,
                                column=column_num).value = cell_num_value

                        if cell_num_value != 0:
                            ws.cell(row=row_num,
                                    column=column_num).fill = redFill

                        else:
                            ws.cell(row=row_num,
                                    column=column_num).fill = greenFill

                        if pattern_flag == 1:
                            pattern_flag = 0
                            first_row = row_num
                            first_col = column_num

                        if 'VDDSA' in source_file_name and 'DUT' in source_file_name:

                            if (row_num - first_row) == def_SDE_index and (column_num - first_col) == def_VDDSA_index:
                                ws.cell(row=row_num,
                                        column=column_num).fill = yellowFill

                        elif 'VDD' in source_file_name and 'DUT' in source_file_name:

                            if (row_num - first_row) == def_SDE_index and (column_num - first_col) == def_VDD_index:
                                ws.cell(row=row_num,
                                        column=column_num).fill = yellowFill

                    else:
                        ws.cell(row=row_num, column=column_num,
                                value=cell_value)

                    column_num = column_num + 1

                row_num = row_num + 1
                VDD_flag = 0
                VDDSA_flag = 0

        sf.close()

for source_file_name in file_list:  # loop the all DUT file

    if 'tb_DLSA_Test_SDE' in source_file_name and 'DUT' not in source_file_name:
        with open(source_file_name, 'r') as sf:
            start_index = source_file_name.index('SDE')
            # generate sheet name
            if get_index(source_file_name, 'VDDSA') != -1:
                end_index = get_index(source_file_name, 'VDDSA')
                sheet_name = source_file_name[start_index: end_index + 5]

            elif get_index(source_file_name, 'VDD') != -1:
                end_index = get_index(source_file_name, 'VDD')
                sheet_name = source_file_name[start_index: end_index + 3]

            ws = wb.create_sheet(sheet_name)

            row_num = 1
            for line_of_sf in sf:

                if 'Vcc' in line_of_sf:
                    VCC_flag = 1

                line_split_list = line_of_sf.rstrip().split(',')
                column_num = 1
                for cell_value in line_split_list:
                    cell_num_value = str_to_num(cell_value)

                    if cell_num_value != -1:
                        ws.cell(row=row_num,
                                column=column_num).value = cell_value

                        if cell_num_value != 0:
                            ws.cell(row=row_num,
                                    column=column_num).fill = redFill

                        else:
                            ws.cell(row=row_num,
                                    column=column_num).fill = greenFill

                        if VCC_flag == 1:
                            VCC_flag = 0
                            first_row = row_num
                            first_col = column_num
                        # fill all DUT's default value with yellow
                        for i in range(0, len(SDE_default_list)):

                            if 'VDDSA' in source_file_name:
                                if (row_num - first_row) == get_index(SDE_DAC, int(SDE_default_list[i], 16)) and (column_num - first_col) == get_index(VDDSA_DAC, int(VDDSA_default_list[i], 16)):
                                    ws.cell(row=row_num,
                                            column=column_num).fill = yellowFill

                            elif 'VDD' in source_file_name:
                                if (row_num - first_row) == get_index(SDE_DAC, int(SDE_default_list[i], 16)) and (column_num - first_col) == get_index(VDD_DAC, int(VDD_default_list[i], 16)):
                                    ws.cell(row=row_num,
                                            column=column_num).fill = yellowFill

                    else:
                        ws.cell(row=row_num, column=column_num,
                                value=cell_value)

                    column_num = column_num + 1
                row_num = row_num + 1
            end_row = row_num
            # insert column and row, then fill the header
            ws.insert_cols(1, amount=1)
            for row_num in range(1, end_row):

                if 'Vcc' in ws.cell(row=row_num, column=2).value:
                    VCC_flag = 1
                cell_num_value = str_to_num(
                    ws.cell(row=row_num, column=2).value)

                if cell_num_value != -1 and VCC_flag == 1:
                    ws.insert_rows(row_num, amount=1)
                    VCC_flag = 0

                    for index in range(0, len(SDE_header)):
                        ws.cell(row=row_num + 1 + index,
                                column=1).value = SDE_header[index]

                    if 'VDDSA' in source_file_name:
                        ws.cell(row=row_num, column=1).value = 'VDDSA='
                        for index in range(0, len(VDDSA_header) - 1):
                            ws.cell(row=row_num, column=index +
                                    2).value = VDDSA_header[index + 1]

                    elif 'VDD' in source_file_name:
                        ws.cell(row=row_num, column=1).value = 'VDD='
                        for index in range(0, len(VDD_header) - 1):
                            ws.cell(row=row_num, column=index +
                                    2).value = VDD_header[index + 1]

# delete the first sheet 'Sheet'
del_sheet = wb['Sheet']
wb.remove(del_sheet)

wb.save('result.xlsx')
print("It's done!")
time.sleep(1)
