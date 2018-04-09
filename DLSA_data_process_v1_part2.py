import os
from openpyxl import Workbook

current_dir = os.getcwd()   # get current directory
file_list = os.listdir(current_dir)  # list file under dir

SDE_DAC = []        # store SDE shmoo value
VDD_DAC = []        # store VDD shmoo value
VDDSA_DAC = []      # store VDDSA shmoo value
SDE_shmoo_step = 2  # SDE shmoo step

for i in range(0x3F, 0x00, -SDE_shmoo_step):
    SDE_DAC.append(i)       # generate SDE_DAC

for i in range(0xF, 0x00, -1):
    VDD_DAC.append(i)
    VDDSA_DAC.append(i)     # generate VDD_DAC and VDDSA_DAC

VDD_DAC.append(0x00)        # add the missing value
VDDSA_DAC.append(0x00)      # add the missing value


def get_index(whole, sub):
    '''get index of a sub string, reture -1 if when whole string didn't include sub string'''
    try:
        index = whole.index(sub)
        return index
    except Exception:
        return -1


def get_Vcc_Pattern_FBC(var_str):
    '''get the value of Vcc or Pattern in the parameter, or get the shmoo points'''
    if 'Vcc' in var_str:
        return 'Vcc', var_str.strip().split()[-1]
    elif 'Pattern' in var_str:
        return 'Pattern', var_str.strip().split()[-1]
    elif 'SDE' in var_str:
        return 'FBC', var_str.strip().split(',')[1:]
    else:
        return False


def cal_margin(var_2d_list, row, col):
    '''calculate the margin pattern dependancy'''
    for r in range(1, min(row, len(var_2d_list) - row - 1) + 1):
        if var_2d_list[row - r][col] != '0' or var_2d_list[row + r][col] != '0':
            r -= 1
            break
    for c in range(1, min(col, len(var_2d_list[0]) - col - 1) + 1):
        if var_2d_list[row][col - c] != '0' or var_2d_list[row][col + c] != '0':
            c -= 1
            break
    return 2*r, c


wb = Workbook()
ws = wb['Sheet']
ws.title = 'Margin_Pattern_Dependancy'

r = 2   # write the dac margin value start from second row of the sheet
for source_file_name in file_list:
    if '.txt' in source_file_name and 'DUT' in source_file_name:
        # initialize these list.
        Vcc = []
        Pattern = []
        FBC = []
        Margin = []
        def_SDE_index = []
        with open(source_file_name, 'r') as sf:
            default_values = sf.readline().rstrip().split(',')  # get default value
            # transform default value to default index in shmoo range
            for value in default_values:

                if 'SDE' in value:
                    SDE_default = value.split('=')[-1]
                    try_SDE_index = get_index(
                        SDE_DAC, int(SDE_default, 16))

                    if try_SDE_index != -1:
                        def_SDE_index.append(try_SDE_index)

                    else:   # if the default value is between to shmoo value
                        def_SDE_index.append(get_index(
                            SDE_DAC, int(SDE_default, 16) - 1))
                        def_SDE_index.append(get_index(
                            SDE_DAC, int(SDE_default, 16) + 1))

                elif 'VDD' in value and 'VDDSA' not in value:
                    VDD_default = value.split('=')[-1]
                    def_VDD_index = get_index(
                        VDD_DAC, int(VDD_default, 16))

                elif 'VDDSA' in value:
                    VDDSA_default = value.split('=')[-1]
                    def_VDDSA_index = get_index(
                        VDDSA_DAC, int(VDDSA_default, 16))

            # calculate the margin and fullfill the Vcc, Pattern, FBC and Margin
            for line in sf:
                get_value = get_Vcc_Pattern_FBC(line)
                if get_value == False:
                    continue
                elif get_value[0] == 'Vcc':
                    Vcc.append(get_value[1])
                    Pattern.append([])
                    Margin.append([])
                elif get_value[0] == 'Pattern':
                    Pattern[len(Vcc)-1].append(get_value[1])
                    Margin[-1].append([])
                    FBC = []
                elif get_value[0] == 'FBC':
                    FBC.append(get_value[1])
                    # when this condition happen, then we can calculate the margin
                    if len(FBC) == len(SDE_DAC):
                        temp_margin = [[], []]
                        for i in range(0, len(def_SDE_index)):
                            for j in range(0, len(temp_margin)):
                                if 'VDDSA' in source_file_name:
                                    temp_margin[j].append(cal_margin(
                                        FBC, def_SDE_index[i], def_VDDSA_index)[j])
                                elif 'VDD' in source_file_name:
                                    temp_margin[j].append(cal_margin(
                                        FBC, def_SDE_index[i], def_VDD_index)[j])
                        # this situation is that default SDE value is between two shmoo point
                        if len(def_SDE_index) == 2:
                            Margin[-1][-1].append(min(temp_margin[0])+1)
                        else:
                            Margin[-1][-1].append(min(temp_margin[0]))
                        Margin[-1][-1].append(min(temp_margin[1]))
        # transport the results to excel
        if get_index(source_file_name, 'DUT') != -1:
            dut_index = get_index(source_file_name, 'DUT')

        for i in range(len(Vcc)):
            ws.cell(row=r, column=1).value = '{},Vcc={}'.format(
                source_file_name[dut_index: dut_index + 4], Vcc[i])
            ws.cell(row=r, column=2).value = 'SDE'
            if 'VDDSA' in source_file_name:
                ws.cell(row=r+1, column=2).value = 'VDDSA'
            elif 'VDD' in source_file_name:
                ws.cell(row=r+1, column=2).value = 'VDD'
            for c in range(len(Margin[0])):
                ws.cell(row=r, column=c+3).value = Margin[i][c][0]
                ws.cell(row=r+1, column=c+3).value = Margin[i][c][1]
            r += 2

for i in range(len(Pattern[0])):
    ws.cell(row=1, column=i+3).value = Pattern[0][i]
wb.save('Margin_Pattern_Dependancy.xlsx')
